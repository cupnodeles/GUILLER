import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border
import msoffcrypto
import re
import io
import zipfile
from datetime import datetime

# ════════════════════════════════════════════════════
# ✅ REQUIRED HEADERS
# ════════════════════════════════════════════════════
REQUIRED_HEADERS = [
    "CUST_ID", "CUST_NAME", "OFC", "HOME", "MOBILE_NO",
    "OTHER CONTACT", "TU", "ADDRESS", "EMAIL", "GENDER",
    "TCL", "OB", "BOS", "AOD", "MAD", "PDA",
    "LAST PAYMENT AMOUNT", "LAST_PAYMENT_DATE", "DELINQUENCY_STRING",
    "ADA_ACCOUNT", "DEBIT_AMOUNT_PREFERENCE", "COLLECTION_CYCLE",
    "UNIT_CODE", "RISK", "AGING", "HO FLAG", "BIRTHDATE",
    "UNIBANKER", "TPAP", "block_code", "MEMO_LINE", "D_CUST_OPN",
    "AREA CODE", "PTP DATE", "CATEGORY/CLASSIF", "LAST DUE DATE",
    "Balance Type", "PTP AMOUNT", "CONTACTED BY", "CLASSIFICATION",
    "TPAP DD", "CLASSIF 2", "INHOUSE", "EMAIL NOTI", "CATEGORY"
]

# ════════════════════════════════════════════════════
# 🔄 CUSTOM COLUMN NAME MAPPING
#    Raw file header → Required header
# ════════════════════════════════════════════════════
CUSTOM_COLUMN_MAP = {
    "OFFICE_PH"        : "OFC",
    "HOME_PH"          : "HOME",
    "LAST_PAYMENT_AMT" : "LAST PAYMENT AMOUNT",
    # ✅ TU → matches directly, no remapping needed
}

# ════════════════════════════════════════════════════
# 📅 DATE COLUMNS TO FORMAT → mm/dd/yyyy
# ════════════════════════════════════════════════════
DATE_COLUMNS = [
    "LAST_PAYMENT_DATE",
    "BIRTHDATE",
    "D_CUST_OPN",
    "LAST DUE DATE",
    "TPAP DD",
    "PTP DATE"
]

# ════════════════════════════════════════════════════
# 📞 PHONE COLUMNS → Fix Scientific Notation
#    + Replace leading "63" with "0"
#    + Add "0" if starts with "9"
# ════════════════════════════════════════════════════
PHONE_COLUMNS = [
    "MOBILE_NO",
    "OFC",
    "HOME"
    # ✅ TU removed — handled by fix_tu() separately
]


# ════════════════════════════════════════════════════
# 📞 Fix Phone Number
#    Step 1: Scientific notation → integer string
#    Step 2: Leading "63"  → "0"
#    Step 3: Leading "9"   → "09"
#    Step 4: Already "0"   → no change
# ════════════════════════════════════════════════════
def fix_phone_number(value):
    if value is None or str(value).strip() in ("", "None"):
        return ""
    try:
        number_str = str(int(float(str(value))))
        if number_str.startswith("63"):
            number_str = "0" + number_str[2:]
        elif number_str.startswith("9"):
            number_str = "0" + number_str
        return number_str
    except Exception:
        clean = str(value).strip()
        if clean.startswith("63"):
            clean = "0" + clean[2:]
        elif clean.startswith("9"):
            clean = "0" + clean
        return clean


# ════════════════════════════════════════════════════
# 📞 Fix TU Column
#    - Split by ";" → process each part
#    - Fix scientific notation
#    - Remove special chars (_, *, spaces)
#    - Fix 639 → 09, 9 → 09
#    - Skip if already starts with "0"
#    - Skip emails and non-phone values
#    - Rejoin with " ; "
# ════════════════════════════════════════════════════
def fix_tu(value):
    if value is None or str(value).strip() in ("", "None"):
        return ""

    text = str(value).strip()

    # ✅ Single scientific notation number (e.g. 6.39098E+11)
    try:
        as_int = str(int(float(text)))
        if as_int.startswith("63"):
            as_int = "0" + as_int[2:]
        elif as_int.startswith("9"):
            as_int = "0" + as_int
        # ✅ Already starts with "0" — no change needed
        return as_int
    except Exception:
        pass

    # ✅ Split by ";" and process each part
    parts  = text.split(";")
    result = []

    for part in parts:
        # ✅ Clean special characters: _, *, spaces
        cleaned = re.sub(r'[_\*\s]', '', part).strip()

        if not cleaned:
            continue

        # ✅ Skip emails
        if "@" in cleaned:
            continue

        # ✅ Try scientific notation conversion
        try:
            as_int = str(int(float(cleaned)))
            if as_int.startswith("63"):
                as_int = "0" + as_int[2:]
            elif as_int.startswith("9"):
                as_int = "0" + as_int
            # ✅ Already starts with "0" — no change needed
            result.append(as_int)
            continue
        except Exception:
            pass

        # ✅ Plain number string — fix prefix only if needed
        if cleaned.startswith("0"):
            pass                        # ✅ Already has 0 — leave it
        elif cleaned.startswith("63"):
            cleaned = "0" + cleaned[2:] # ✅ 639 → 09
        elif cleaned.startswith("9"):
            cleaned = "0" + cleaned     # ✅ 9 → 09

        # ✅ Only keep if it looks like a phone number (digits only, 7-13 digits)
        if cleaned.isdigit() and 7 <= len(cleaned) <= 13:
            result.append(cleaned)

    return " ; ".join(result)


# ════════════════════════════════════════════════════
# 📅 Format Date → mm/dd/yyyy
# ════════════════════════════════════════════════════
def format_date(value):
    if value is None or str(value).strip() in ("", "None"):
        return ""
    try:
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%Y")
        if hasattr(value, "strftime"):
            return value.strftime("%m/%d/%Y")
        text   = str(value).strip()
        parsed = datetime.strptime(text.split(" ")[0], "%Y-%m-%d")
        return parsed.strftime("%m/%d/%Y")
    except Exception:
        return str(value)


# ════════════════════════════════════════════════════
# 🔑 Detect Cycle → Password
# ════════════════════════════════════════════════════
def get_password_from_filename(filename):
    match = re.search(r'cycle[_\s]?(\d+)', filename, re.IGNORECASE)
    if match:
        cycle_number = match.group(1)
        return f"CYCLE_{cycle_number}*", cycle_number
    return None, None


# ════════════════════════════════════════════════════
# 📁 Build Output Filename
# ════════════════════════════════════════════════════
def build_output_filename(cycle_number, date=None):
    if date is None:
        date = datetime.today()
    return f"BPI_XDAYS_C{cycle_number}_HEADER_ALIGNED_{date.strftime('%m-%d-%Y')}.xlsx"


# ════════════════════════════════════════════════════
# 🔓 Unlock Excel
# ════════════════════════════════════════════════════
def unlock_excel(uploaded_file, password):
    try:
        decrypted   = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(uploaded_file)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted, None
    except Exception as e:
        return None, str(e)


# ════════════════════════════════════════════════════
# 📊 Generate Cycle Summary (Count & Sum of OB)
# ════════════════════════════════════════════════════
def generate_cycle_summary(processed_files):
    summary = []
    for output_filename, cycle_number, file_bytes in processed_files:
        file_bytes.seek(0)
        wb  = openpyxl.load_workbook(file_bytes)
        ws  = wb["WL"]

        # ── Get headers from row 1 ──
        headers = [cell.value for cell in ws[1]]
        headers = [str(h).strip().upper() if h else "" for h in headers]

        # ── Find OB column index ──
        ob_idx = next(
            (i for i, h in enumerate(headers) if h == "OB"),
            None
        )

        # ── Count rows & sum OB ──
        account_count = 0
        ob_sum        = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # ✅ Only count rows with data in Column A (CUST_ID)
            cust_id = row[0]
            if cust_id is None or str(cust_id).strip() in ("", "None"):
                continue
            account_count += 1

            # ✅ Sum OB column
            if ob_idx is not None and ob_idx < len(row):
                ob_val = row[ob_idx]
                try:
                    ob_sum += float(ob_val)
                except (TypeError, ValueError):
                    pass

        summary.append({
            "Cycle"          : f"Cycle {cycle_number}",
            "Output File"    : output_filename,
            "Total Accounts" : account_count,
            "Total OB"       : round(ob_sum, 2),
        })

        file_bytes.seek(0)  # ✅ Reset after reading

    return pd.DataFrame(summary)


# ════════════════════════════════════════════════════
# 📐 Align & Format Headers
# ════════════════════════════════════════════════════
def align_and_format_headers(file_bytes, cycle_number=None):
    wb = openpyxl.load_workbook(file_bytes)

    # ── Step 1: Find WL Sheet ──
    wl_sheet = None
    for sheet_name in wb.sheetnames:
        if "WL" in sheet_name.upper():
            wl_sheet = wb[sheet_name]
            break

    if wl_sheet is None:
        raise ValueError(
            f"No WL sheet found! "
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )

    ws = wl_sheet

    # ── Step 2: Read Existing Headers ──
    existing_headers = [cell.value for cell in ws[1]]

    def normalize(val):
        return str(val).strip().upper().replace("_", " ") if val else ""

    # ── Step 3: Apply Custom Column Map ──
    remapped_headers = []
    for h in existing_headers:
        raw_upper = str(h).strip().upper() if h else ""
        matched   = next(
            (v for k, v in CUSTOM_COLUMN_MAP.items()
             if k.upper() == raw_upper),
            None
        )
        remapped_headers.append(matched if matched else h)

    remapped_normalized = [normalize(h) for h in remapped_headers]

    # ── Required headers & normalized versions ──
    required_for_mapping = REQUIRED_HEADERS
    required_normalized  = [normalize(h) for h in required_for_mapping]

    # ── Normalized sets for quick lookup ──
    date_cols_normalized  = [normalize(d) for d in DATE_COLUMNS]
    phone_cols_normalized = [normalize(p) for p in PHONE_COLUMNS]

    # ── Step 4: Find COLLECTION_CYCLE index first ──
    collection_cycle_idx = next(
        (i for i, h in enumerate(required_for_mapping)
         if normalize(h) == "COLLECTION CYCLE"),
        None
    )

    # ── Step 5: Map Columns ──
    #    ✅ Skip COLLECTION_CYCLE — always injected from filename
    col_mapping = {}
    for req_idx, req_norm in enumerate(required_normalized):
        if req_idx == collection_cycle_idx:
            continue
        for ex_idx, ex_norm in enumerate(remapped_normalized):
            if req_norm == ex_norm:
                col_mapping[req_idx] = ex_idx
                break

    # ── Step 6: Read All Data ──
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]

    # ── Step 7: Reorder + Format ──
    new_data = []
    for row in data:

        # ✅ Skip rows with no CUST_ID in Column A
        cust_id_idx = col_mapping.get(0)
        cust_id_val = (
            row[cust_id_idx]
            if cust_id_idx is not None and cust_id_idx < len(row)
            else None
        )
        if cust_id_val is None or str(cust_id_val).strip() in ("", "None"):
            continue

        new_row = []
        for req_idx in range(len(required_for_mapping)):

            # ✅ COLLECTION_CYCLE → always injected from filename
            if req_idx == collection_cycle_idx:
                new_row.append(
                    str(cycle_number).zfill(2) if cycle_number else None
                )
                continue

            if req_idx in col_mapping:
                ex_idx     = col_mapping[req_idx]
                cell_value = row[ex_idx] if ex_idx < len(row) else None
                col_name   = normalize(required_for_mapping[req_idx])
                is_empty   = (
                    cell_value is None or
                    str(cell_value).strip() in ("", "None")
                )

                # ✅ DATE COLUMNS → mm/dd/yyyy
                if col_name in date_cols_normalized:
                    new_row.append(
                        "" if is_empty else format_date(cell_value)
                    )

                # ✅ PHONE COLUMNS → fix sci notation + 63→0 + 9→09
                elif col_name in phone_cols_normalized:
                    new_row.append(
                        "" if is_empty else fix_phone_number(cell_value)
                    )

                # ✅ TU COLUMN → dedicated fix for mixed content
                elif col_name == "TU":
                    new_row.append(
                        "" if is_empty else fix_tu(cell_value)
                    )

                else:
                    new_row.append(cell_value)
            else:
                new_row.append(None)

        new_data.append(new_row)

    # ── Step 8: Delete ALL Sheets ──
    for sheet in wb.sheetnames:
        del wb[sheet]

    # ── Step 9: Create New Clean Sheet ──
    ws_new = wb.create_sheet(title="WL")

    # ════════════════════════════════════════════════════
    # 🎨 STYLES
    # ════════════════════════════════════════════════════
    header_font = Font(
        name  = "Calibri",
        bold  = True,
        color = "000000",
        size  = 11
    )
    header_alignment = Alignment(
        horizontal = "center",
        vertical   = "center",
        wrap_text  = False
    )
    data_font = Font(
        name  = "Calibri",
        bold  = False,
        color = "000000",
        size  = 11
    )
    data_alignment = Alignment(
        horizontal = "left",
        vertical   = "center"
    )

    # ── Step 10: Write Headers ──
    for col_idx, header in enumerate(REQUIRED_HEADERS, start=1):
        cell           = ws_new.cell(row=1, column=col_idx, value=header.upper())
        cell.font      = header_font
        cell.alignment = header_alignment
        cell.fill      = PatternFill(fill_type=None)
        cell.border    = Border()

    # ── Step 11: Write Data Rows ──
    for row_idx, row_data in enumerate(new_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws_new.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = data_font
            cell.alignment = data_alignment
            cell.fill      = PatternFill(fill_type=None)
            cell.border    = Border()

    # ── Step 12: Auto-fit Column Widths ──
    for col in ws_new.columns:
        max_len    = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws_new.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # ── Step 13: Row Height & Freeze ──
    ws_new.row_dimensions[1].height = 20
    ws_new.freeze_panes = "A2"

    # ── Step 14: Save ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, col_mapping, existing_headers, wl_sheet.title


# ════════════════════════════════════════════════════
# ⚙️ Process Single File
# ════════════════════════════════════════════════════
def process_file(uploaded_file, custom_date=None):
    filename = uploaded_file.name

    password, cycle_number = get_password_from_filename(filename)
    if not password:
        return None, None, None, None, None, None, "Cycle number not detected."

    decrypted_file, error = unlock_excel(uploaded_file, password)
    if error:
        return None, None, cycle_number, None, None, None, error

    try:
        formatted_file, col_mapping, existing_headers, wl_sheet_name = \
            align_and_format_headers(decrypted_file, cycle_number)
    except ValueError as e:
        return None, None, cycle_number, None, None, None, str(e)

    output_filename = build_output_filename(cycle_number, custom_date)

    return (
        formatted_file, output_filename, cycle_number,
        col_mapping, existing_headers, wl_sheet_name, None
    )


# ════════════════════════════════════════════════════
# 🗜️ Create ZIP
# ════════════════════════════════════════════════════
def create_zip(processed_files, today_str):
    zip_buffer  = io.BytesIO()
    root_folder = f"BPI_XDAYS_HEADER_ALIGNED_{today_str}"
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for output_filename, cycle_number, file_bytes in processed_files:
            file_bytes.seek(0)
            zf.writestr(
                f"{root_folder}/{output_filename}",
                file_bytes.read()
            )
    zip_buffer.seek(0)
    return zip_buffer


# ════════════════════════════════════════════════════
# 🖥️ STREAMLIT UI
# ════════════════════════════════════════════════════
st.set_page_config(
    page_title = "Batch Upload",
    page_icon  = "📁",
    layout     = "wide"
)

st.title("📁 Batch File Upload")
st.markdown(
    "Upload **multiple password-protected Excel files** at once. "
    "App will extract the **WL sheet**, align headers, "
    "format **dates → mm/dd/yyyy**, fix **phone numbers**, "
    "and package into **1 ZIP folder**."
)
st.divider()

# ── Date Selector ──
st.markdown("### 📅 Output File Date")
date_option = st.radio(
    "Date to use in output filenames:",
    ["📅 Use Today's Date (Auto)", "✏️ Pick a Custom Date"],
    horizontal=True
)
custom_date = None
if date_option == "✏️ Pick a Custom Date":
    picked_date = st.date_input("Select date:", value=datetime.today())
    custom_date = datetime.combine(picked_date, datetime.min.time())
    today_str   = picked_date.strftime("%m-%d-%Y")
    st.success(f"📅 Using custom date: **{today_str}**")
else:
    today_str = datetime.today().strftime("%m-%d-%Y")
    st.success(f"📅 Using today's date: **{today_str}**")

st.divider()

# ── Batch Uploader ──
uploaded_files = st.file_uploader(
    "📤 Upload Excel Files (.xlsx) — Select multiple files at once!",
    type                  = ["xlsx"],
    accept_multiple_files = True
)

if uploaded_files:
    total = len(uploaded_files)

    # ── File Queue Preview ──
    st.markdown(f"### 📁 {total} File(s) Queued for Processing")
    with st.expander("👁️ View File Queue", expanded=True):
        h1, h2, h3, h4 = st.columns([3, 1, 1, 4])
        h1.markdown("**📄 Original Filename**")
        h2.markdown("**🔢 Cycle**")
        h3.markdown("**🔑 Password**")
        h4.markdown("**📄 Output Filename**")
        st.divider()
        for f in uploaded_files:
            _, cycle = get_password_from_filename(f.name)
            out_name = (
                build_output_filename(cycle, custom_date)
                if cycle else "❌ Cycle not detected"
            )
            c1, c2, c3, c4 = st.columns([3, 1, 1, 4])
            c1.markdown(f"`{f.name}`")
            c2.markdown(f"**{cycle}**" if cycle else "❌")
            c3.markdown(f"`CYCLE_{cycle}*`" if cycle else "❌")
            c4.markdown(f"`{out_name}`")

    st.divider()

    # ── Process Button ──
    if st.button(
        f"⚙️ Process All {total} File(s)",
        use_container_width = True,
        type                = "primary"
    ):
        processed_files = []
        success_count   = 0
        fail_count      = 0
        results_summary = []

        progress_bar = st.progress(0, text="Starting batch processing...")
        status_area  = st.empty()

        for i, uploaded_file in enumerate(uploaded_files):
            filename = uploaded_file.name
            status_area.markdown(
                f"⏳ Processing **{filename}** ({i+1}/{total})..."
            )

            (
                formatted_file, output_filename, cycle_number,
                col_mapping, existing_headers, wl_sheet_name, error
            ) = process_file(uploaded_file, custom_date)

            if error:
                fail_count += 1
                results_summary.append({
                    "Original File"  : filename,
                    "Cycle"          : cycle_number or "N/A",
                    "Password Used"  : f"CYCLE_{cycle_number}*" if cycle_number else "N/A",
                    "WL Sheet"       : "❌ Not Found",
                    "Output Filename": "— Failed —",
                    "Status"         : f"❌ {error}",
                    "Matched Cols"   : "—",
                    "Missing Cols"   : "—",
                })
            else:
                success_count += 1
                matched = len(col_mapping)
                missing = len(REQUIRED_HEADERS) - matched
                processed_files.append(
                    (output_filename, cycle_number, formatted_file)
                )
                results_summary.append({
                    "Original File"  : filename,
                    "Cycle"          : cycle_number,
                    "Password Used"  : f"CYCLE_{cycle_number}*",
                    "WL Sheet"       : f"✅ {wl_sheet_name}",
                    "Output Filename": output_filename,
                    "Status"         : "✅ Success",
                    "Matched Cols"   : matched,
                    "Missing Cols"   : missing,
                })

            progress_bar.progress(
                int((i + 1) / total * 100),
                text=f"Processing {i+1}/{total} files..."
            )

        progress_bar.progress(100, text="✅ All files processed!")
        status_area.empty()
        st.divider()

        # ── Batch Processing Summary ──
        st.markdown("## 📊 Batch Processing Summary")
        m1, m2, m3 = st.columns(3)
        m1.metric("📁 Total Files",  total)
        m2.metric("✅ Successful",    success_count)
        m3.metric("❌ Failed",        fail_count)
        st.divider()

        df_summary = pd.DataFrame(results_summary)
        st.dataframe(df_summary, use_container_width=True, hide_index=True)
        st.divider()

        # ── ✅ Cycle Summary — Account Count & OB Sum ──
        if processed_files:
            st.markdown("## 📊 Cycle Summary — Accounts & OB")
            df_cycle_summary = generate_cycle_summary(processed_files)

            # ── Metrics per Cycle ──
            cols = st.columns(len(df_cycle_summary))
            for col, row in zip(cols, df_cycle_summary.itertuples()):
                col.metric(
                    label = f"🔢 {row.Cycle}",
                    value = f"{row._3:,} accounts",
                    delta = f"OB: {row._4:,.2f}"
                )

            st.divider()

            # ── Full Summary Table ──
            st.dataframe(
                df_cycle_summary.style.format({
                    "Total Accounts" : "{:,}",
                    "Total OB"       : "{:,.2f}"
                }),
                use_container_width = True,
                hide_index          = True
            )
            st.divider()

        if processed_files:

            # ── Individual Downloads ──
            with st.expander("📄 Download Files Individually", expanded=False):
                for out_fname, cycle_num, file_bytes in sorted(
                    processed_files, key=lambda x: int(x[1])
                ):
                    file_bytes.seek(0)
                    st.download_button(
                        label               = f"📥 [Cycle {cycle_num}]  {out_fname}",
                        data                = file_bytes,
                        file_name           = out_fname,
                        mime                = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width = True,
                        key                 = f"dl_{out_fname}"
                    )

            # ── ZIP Download ──
            st.markdown("### 📦 Download All Files as ZIP")
            zip_file     = create_zip(processed_files, today_str)
            zip_filename = f"BPI_XDAYS_HEADER_ALIGNED_{today_str}.zip"

            st.download_button(
                label               = f"📦 Download ZIP  →  {zip_filename}",
                data                = zip_file,
                file_name           = zip_filename,
                mime                = "application/zip",
                use_container_width = True,
                type                = "primary"
            )

            st.success(
                f"🎉 **{success_count} file(s) ready!** "
                f"Each file has **1 clean WL sheet** with "
                f"**{len(REQUIRED_HEADERS)} CAPS + BOLD headers**, "
                f"**dates → mm/dd/yyyy**, "
                f"**phone numbers fixed (63→0, 9→09)**, "
                f"and **TU column cleaned**. 🗂️"
            )

        if fail_count > 0:
            st.warning(
                f"⚠️ **{fail_count} file(s) failed.**\n\n"
                "Possible reasons:\n"
                "- Filename missing cycle number\n"
                "- Wrong password\n"
                "- No WL sheet found in file\n\n"
                "Example: `XDAYS Cycle 9 WORKLIST 06262025.xlsx`"
            )